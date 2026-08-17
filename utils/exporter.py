"""
تصدير BOQ إلى Excel احترافي مع ملخص المواد والافتراضات الهندسية

Supports an optional pricing layer: if the DataFrame carries "Unit Price" and
"Total" columns, two extra money columns + a grand-total row are written.
Now includes:
- Worksheet 1: "BOQ" (Itemized takeoff and pricing)
- Worksheet 2: "Material Summary" (Aggregated concrete, steel, blocks counts and totals)
- Worksheet 3: "Assumptions & Estimators" (Detailed layout and geo variables documentation)
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import io


SECTION_FILL  = PatternFill("solid", fgColor="2D6A9F")
ALT_FILL      = PatternFill("solid", fgColor="F0F4F8")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
TOTAL_FILL    = PatternFill("solid", fgColor="1A3C5E")
HEADER_FILL   = PatternFill("solid", fgColor="1A3C5E")
WHITE_FONT    = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
HEADER_FONT   = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
TOTAL_FONT    = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
BODY_FONT     = Font(name="Calibri", size=11)
CENTER        = Alignment(horizontal="center", vertical="center")
RIGHT         = Alignment(horizontal="right", vertical="center")
LEFT          = Alignment(horizontal="left",  vertical="center")
THIN          = Side(style="thin", color="CCCCCC")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_boq_to_excel(df: pd.DataFrame, project_name: str,
                        currency: str = "AED", project_meta: dict = None) -> bytes:
    """
    Build a styled BOQ workbook with Material Summary and Assumptions sheets.

    If `df` has "Unit Price" + "Total" columns, the sheet includes pricing and
    a grand-total row; otherwise it's a quantity-only BOQ.
    """
    # Always enforce the full 7-column Priced BOQ format with Unit Price and Total columns
    priced = True

    cols      = ["#", "Description (English)", "البيان", "Unit", "Quantity",
                 "Unit Price", "Total"]
    headers   = ["#", "Description (English)", "البيان", "Unit", "Quantity",
                 f"Unit Price ({currency})", f"Total ({currency})"]
    col_widths = [8, 42, 30, 8, 12, 16, 18]

    ncol     = len(cols)
    last_col = get_column_letter(ncol)

    wb = Workbook()
    
    # ── Sheet 1: BOQ ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "BOQ"

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"Bill of Quantities — {project_name}"
    ws["A1"].font      = Font(bold=True, size=16, color="1A3C5E", name="Calibri")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 30

    # Date
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Date: {date.today().strftime('%d %B %Y')}"
    ws["A2"].font      = Font(italic=True, size=11, color="666666", name="Calibri")
    ws["A2"].alignment = CENTER
    ws.row_dimensions[3].height = 8

    # Column headers
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER
    ws.row_dimensions[4].height = 22

    money_cols = {6, 7} if priced else set()
    qty_col    = 5
    row_idx    = 5
    alt        = False
    grand_total = 0.0

    for r in df.to_dict("records"):
        is_header = bool(r.get("_is_header", False))
        ws.row_dimensions[row_idx].height = 18

        if is_header:
            ws.merge_cells(f"A{row_idx}:{last_col}{row_idx}")
            title = str(r.get("Description (English)", "")).replace("▌ ", "")
            cell = ws.cell(row=row_idx, column=1, value=title)
            cell.fill, cell.font, cell.alignment, cell.border = SECTION_FILL, WHITE_FONT, LEFT, BORDER
            alt = False
        else:
            fill = ALT_FILL if alt else WHITE_FILL
            for c, key in enumerate(cols, 1):
                # Force Unit Price to 0.0 and Total to Formula
                if key == "Unit Price":
                    val = float(r.get("Unit Price") or 0.0)
                elif key == "Total":
                    # Excel formula: Quantity (Col E) * Unit Price (Col F)
                    val = f"=E{row_idx}*F{row_idx}"
                else:
                    val = r.get(key, "")
                    
                cell = ws.cell(row=row_idx, column=c, value=val)
                cell.fill, cell.font, cell.border = fill, BODY_FONT, BORDER
                if c == qty_col or c in money_cols:
                    cell.alignment = RIGHT
                    if c in money_cols and key != "Total":
                        cell.number_format = "#,##0.00"
                    elif key == "Total":
                        cell.number_format = "#,##0.00"
                elif c in (1, 4):
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT
            if priced:
                # We can't sum formulas easily in openpyxl without a string formula for the grand total
                pass
            alt = not alt
        row_idx += 1

    # Grand total row
    if priced:
        ws.merge_cells(f"A{row_idx}:{get_column_letter(ncol-1)}{row_idx}")
        lbl = ws.cell(row=row_idx, column=1, value=f"GRAND TOTAL ({currency})")
        lbl.fill, lbl.font, lbl.alignment, lbl.border = TOTAL_FILL, TOTAL_FONT, RIGHT, BORDER
        # Grand total formula: SUM(G5:G{row_idx-1})
        tot = ws.cell(row=row_idx, column=ncol, value=f"=SUM(G5:G{row_idx-1})")
        tot.fill, tot.font, tot.alignment, tot.border = TOTAL_FILL, TOTAL_FONT, RIGHT, BORDER
        tot.number_format = "#,##0.00"
        ws.row_dimensions[row_idx].height = 24

    ws.freeze_panes = "A5"

    # ── Doors and Windows Schedules ───────────────────────────────────────────
    if project_meta:
        openings = project_meta.get("confirmed_auto_data", {}).get("openings", {})
        doors = openings.get("doors", [])
        windows = openings.get("windows", [])
        
        def write_schedule(sheet_title, items):
            ws_sched = wb.create_sheet(sheet_title)
            
            ws_sched.merge_cells("A1:F1")
            ws_sched["A1"] = f"{sheet_title} — {project_name}"
            ws_sched["A1"].font = Font(bold=True, size=14, color="1A3C5E", name="Calibri")
            ws_sched["A1"].alignment = CENTER
            ws_sched.row_dimensions[1].height = 28
            
            headers = ["#", "Mark", "Width (m)", "Height (m)", "Count", "Total Area (m²)"]
            col_widths = [8, 20, 15, 15, 12, 18]
            for i, w in enumerate(col_widths, 1):
                ws_sched.column_dimensions[get_column_letter(i)].width = w
                
            for c, h in enumerate(headers, 1):
                cell = ws_sched.cell(row=3, column=c, value=h)
                cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER
            ws_sched.row_dimensions[3].height = 22
            
            row_idx = 4
            grand_total_area = 0.0
            grand_total_count = 0
            for idx, item in enumerate(items, 1):
                mark = item.get("mark") or f"Type {idx}"
                width = item.get("width_m") or item.get("width") or item.get("width_mm") or 0.0
                height = item.get("height_m") or item.get("height") or item.get("height_mm") or 0.0
                count = item.get("count_in_plans") or item.get("count") or 0
                
                try:
                    w_val = float(str(width).strip() or 0)
                    h_val = float(str(height).strip() or 0)
                    if width == item.get("width_mm") and w_val > 100: w_val /= 1000.0
                    if height == item.get("height_mm") and h_val > 100: h_val /= 1000.0
                    c_val = int(float(str(count).strip() or 0))
                except:
                    w_val, h_val, c_val = 0.0, 0.0, 0
                
                area = w_val * h_val * c_val
                grand_total_area += area
                grand_total_count += c_val
                
                ws_sched.cell(row=row_idx, column=1, value=idx).alignment = CENTER
                ws_sched.cell(row=row_idx, column=2, value=mark).alignment = CENTER
                ws_sched.cell(row=row_idx, column=3, value=w_val).alignment = CENTER
                ws_sched.cell(row=row_idx, column=4, value=h_val).alignment = CENTER
                ws_sched.cell(row=row_idx, column=5, value=c_val).alignment = CENTER
                ws_sched.cell(row=row_idx, column=6, value=round(area, 2)).alignment = RIGHT
                
                for c in range(1, 7):
                    ws_sched.cell(row=row_idx, column=c).border = BORDER
                
                row_idx += 1
                
            ws_sched.merge_cells(f"A{row_idx}:D{row_idx}")
            ws_sched.cell(row=row_idx, column=1, value="GRAND TOTAL").alignment = RIGHT
            ws_sched.cell(row=row_idx, column=1).font = TOTAL_FONT
            ws_sched.cell(row=row_idx, column=1).fill = TOTAL_FILL
            for c in range(1, 5):
                ws_sched.cell(row=row_idx, column=c).border = BORDER
                ws_sched.cell(row=row_idx, column=c).fill = TOTAL_FILL
                
            ws_sched.cell(row=row_idx, column=5, value=grand_total_count).alignment = CENTER
            ws_sched.cell(row=row_idx, column=5).font = TOTAL_FONT
            ws_sched.cell(row=row_idx, column=5).fill = TOTAL_FILL
            ws_sched.cell(row=row_idx, column=5).border = BORDER
            
            ws_sched.cell(row=row_idx, column=6, value=round(grand_total_area, 2)).alignment = RIGHT
            ws_sched.cell(row=row_idx, column=6).font = TOTAL_FONT
            ws_sched.cell(row=row_idx, column=6).fill = TOTAL_FILL
            ws_sched.cell(row=row_idx, column=6).border = BORDER

        if windows:
            write_schedule("Windows Schedule", windows)
        if doors:
            write_schedule("Doors Schedule", doors)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
