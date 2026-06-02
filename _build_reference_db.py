"""
Builds reference_ranges.json from a folder of training BOQ files.

Usage:
    # Full rebuild from default folder
    py _build_reference_db.py

    # Rebuild from custom folder
    py _build_reference_db.py --base "B:\\work\\new_villas"

    # Incrementally add a single new project (doesn't touch existing data)
    py _build_reference_db.py --add "B:\\work\\new_villas\\villa_42" \\
        --villa-type G+1 --floors 3 --plot-area 750 --gf-area 330

Each project gets metadata so the validator can filter "similar villas only":
    villa_type:  G | G+1 | G+2 | G+B+1 ...
    floors:      total structural levels (incl. roof)
    plot_area:   plot area m²
    gf_area:     GF built-up area m²

Output JSON layout:
{
  "<item_key>": {
    "unit": "m3",
    "count": 14,
    "min": 1.2, "max": 8.4, "median": 3.6, "mean": 3.8, "stdev": 1.4,
    "values": [...],
    "by_villa_type": {
      "G+1": {"count": 8, "median": 3.2},
      "G+2": {"count": 4, "median": 5.1}
    }
  },
  ...
}

There's also a top-level "_projects" key listing every project with metadata,
so the validator can find sibling projects similar to the one being analyzed.
"""
import openpyxl, os, json, re, warnings, statistics, argparse, sys
warnings.filterwarnings('ignore')

BASE = r'B:\work\trainning'

# ── Map item description keywords → our system key ────────────────────────────
KEYWORD_MAP = [
    # sub-structure — ORDER MATTERS: specific rules before generic ones
    ("excavat",                          "excavation"),
    ("excvat",                           "excavation"),    # common BOQ typo
    ("road base",                        "road_base"),
    ("roadbase",                         "road_base"),
    ("back fill",                        "backfill"),
    ("backfill",                         "backfill"),
    # PCC/blinding — must come BEFORE generic "footing"
    ("pcc",                              "foundation_pcc"),
    ("ppc",                              "foundation_pcc"),  # BOQ typo for PCC
    ("plain cement",                     "foundation_pcc"),
    ("blinding",                         "foundation_pcc"),
    ("100mm thick",                      "foundation_pcc"),  # "100mm thick under footing"
    # RCC footings — specific "r.c.c" or "rcc" + footing = total footing concrete
    ("r.c.c foot",                       "foundation_concrete"),
    ("rcc foot",                         "foundation_concrete"),
    ("r.c.c. foot",                      "foundation_concrete"),
    # wall footing is a sub-item — keep separate, don't merge with RCC footings
    ("wall footing",                     "wall_footing"),
    ("strip footing",                    "wall_footing"),
    # generic "footing" fallback — only if nothing above matched
    ("footing",                          "foundation_concrete"),
    ("anti-termite",                     "anti_termite"),
    ("anti termaite",                    "anti_termite"),
    ("poly sheet",                       "polyethylene"),
    ("polythene",                        "polyethylene"),
    ("neck col",                         "neck_column_concrete"),
    # block-work rules MUST come before "tie beam" to avoid "solid block under tie beams" being
    # classified as tie_beam_concrete
    ("thermal block",                    "thermal_block"),
    ("hollow block",                     "internal_block"),
    ("solid block",                      "solid_block"),
    ("block work",                       "solid_block"),
    # tie beam concrete — only plain concrete items should reach this rule
    ("tie beam",                         "tie_beam_concrete"),
    ("strap beam",                       "tie_beam_concrete"),
    ("slab on grade",                    "slab_on_grade"),
    ("ground slab",                      "slab_on_grade"),
    # super-structure
    ("rcc col",                          "column_concrete"),
    ("column",                           "column_concrete"),
    ("rcc beam",                         "beam_concrete"),
    ("beam concrete",                    "beam_concrete"),
    ("rcc slab",                         "slab_concrete"),
    ("slab concrete",                    "slab_concrete"),
    ("1st floor slab",                   "slab_concrete_1st"),
    ("2nd floor slab",                   "slab_concrete_2nd"),
    ("roof slab",                        "slab_concrete_roof"),
    ("staircase",                        "staircase_concrete"),
    ("parapet",                          "parapet"),
    # block work (remaining types not already listed above)
    ("25cm",                             "thermal_block"),
    ("20cm block",                       "internal_block"),
    ("internal block",                   "internal_block"),
    # finishes
    ("external plaster",                 "external_plaster"),
    ("ext. plaster",                     "external_plaster"),
    ("internal plaster",                 "internal_plaster"),
    ("int. plaster",                     "internal_plaster"),
    ("external paint",                   "external_paint"),
    ("internal paint",                   "internal_paint"),
    ("int paint",                        "internal_paint"),
    ("ceramic tile",                     "ceramic_tiles"),
    ("porcelain tile",                   "ceramic_tiles"),
    ("floor tile",                       "dry_floor"),
    ("wall tile",                        "wall_tiles"),
    ("marble",                           "marble"),
    ("granite",                          "granite"),
    ("skirting",                         "skirting"),
    # waterproofing
    ("waterproof",                       "waterproofing"),
    ("bitumen coat",                     "bitumen"),
    # openings
    ("aluminium door",                   "door_openings"),
    ("wooden door",                      "door_openings"),
    ("door",                             "door_openings"),
    ("window",                           "window_openings"),
    # external
    ("interlock",                        "interlock_paving"),
    ("compound wall",                    "compound_wall"),
    ("boundary wall",                    "compound_wall"),
]

SKIP_UNITS = {'l.s', 'ls', 'l/s', 'lump', '%', 'ps', 'p.s'}

def match_item_key(desc: str) -> str | None:
    d = desc.lower()
    for kw, key in KEYWORD_MAP:
        if kw in d:
            return key
    return None

def parse_qty(val) -> float | None:
    if val is None:
        return None
    try:
        f = float(str(val).replace(',', '').strip())
        if f > 0:
            return f
    except:
        pass
    return None

def parse_unit(val) -> str:
    if val is None:
        return ''
    u = str(val).strip().lower()
    return u

# ── Parse a single project folder ──────────────────────────────────────────────
def parse_project_folder(proj_path: str, proj_name: str) -> list[dict]:
    """Returns list of {item_key, qty, unit, desc} for one project."""
    entries = []
    if not os.path.isdir(proj_path):
        return entries
    for fname in os.listdir(proj_path):
        if not fname.endswith('.xlsx'):
            continue
        fpath = os.path.join(proj_path, fname)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            for sh in wb.sheetnames:
                ws = wb[sh]
                for row in ws.iter_rows(values_only=True):
                    desc = str(row[1] or '').strip() if len(row) > 1 else ''
                    if not desc or len(desc) < 4:
                        continue
                    qty_val, unit_val = None, ''
                    for i, cell in enumerate(row):
                        cu = parse_unit(cell)
                        if cu in ('m3','m²','m³','m2','sqm','rm','nr','no','lm') and cu not in SKIP_UNITS:
                            unit_val = cu
                            if i > 0:
                                qty_val = parse_qty(row[i-1])
                        if cu == 'qty.':
                            qty_val = parse_qty(row[i+1]) if i+1 < len(row) else None
                    if len(row) >= 8 and qty_val is None:
                        qty_val  = parse_qty(row[6])
                        unit_val = parse_unit(row[7])
                    if qty_val is None or unit_val in SKIP_UNITS or unit_val == '':
                        continue
                    key = match_item_key(desc)
                    if key is None:
                        continue
                    entries.append({
                        "item_key": key,
                        "qty":      qty_val,
                        "unit":     unit_val,
                        "project":  proj_name,
                        "desc":     desc[:60],
                    })
        except Exception as e:
            print(f"  SKIP {fpath}: {e}")
    return entries


def _aggregate(entries_by_key: dict, projects_meta: dict, out_path: str):
    """Build reference_ranges.json from a {item_key → list_of_entries} dict."""
    reference = {}
    for key, entries in entries_by_key.items():
        qtys = [e["qty"] for e in entries]
        if len(qtys) < 2:
            continue
        unit  = entries[0]["unit"]
        # by_villa_type breakdown for smarter validation
        by_vt: dict[str, list[float]] = {}
        for e in entries:
            vt = projects_meta.get(e["project"], {}).get("villa_type", "unknown")
            by_vt.setdefault(vt, []).append(e["qty"])
        reference[key] = {
            "unit":   unit,
            "count":  len(qtys),
            "min":    round(min(qtys), 2),
            "max":    round(max(qtys), 2),
            "median": round(statistics.median(qtys), 2),
            "mean":   round(statistics.mean(qtys), 2),
            "stdev":  round(statistics.stdev(qtys), 2) if len(qtys) > 2 else 0,
            "values": sorted(qtys),
            "samples": [e["desc"] for e in entries[:3]],
            "by_villa_type": {
                vt: {
                    "count":  len(vals),
                    "median": round(statistics.median(vals), 2),
                    "min":    round(min(vals), 2),
                    "max":    round(max(vals), 2),
                }
                for vt, vals in by_vt.items() if len(vals) >= 2
            },
        }
    reference["_projects"] = projects_meta
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(reference, f, indent=2, ensure_ascii=False)
    return reference


def _print_summary(reference: dict, out_path: str):
    items = {k: v for k, v in reference.items() if not k.startswith("_")}
    print(f"\nOK Reference DB built -> {out_path}")
    print(f"   Items with ranges: {len(items)}")
    print(f"   Projects:          {len(reference.get('_projects', {}))}")
    for k, v in sorted(items.items()):
        vt_summary = ""
        if v.get("by_villa_type"):
            vt_summary = "  " + " ".join(
                f"[{vt} n={d['count']}]" for vt, d in v["by_villa_type"].items()
            )
        print(f"   {k:30s} {v['count']:2d}p  "
              f"{v['min']:8.1f} - {v['max']:8.1f} {v['unit']:3s}  "
              f"(med {v['median']:.1f}){vt_summary}")


# ── Modes ──────────────────────────────────────────────────────────────────────
def mode_rebuild(base: str, out_path: str):
    """Full rebuild from a folder of project subfolders."""
    entries_by_key = {}
    projects_meta  = {}
    # Try to load existing project metadata so we keep it across rebuilds
    if os.path.exists(out_path):
        try:
            with open(out_path, encoding='utf-8') as f:
                old = json.load(f)
            projects_meta = old.get("_projects", {})
        except Exception:
            pass

    for proj in sorted(os.listdir(base)):
        proj_path = os.path.join(base, proj)
        if not os.path.isdir(proj_path):
            continue
        es = parse_project_folder(proj_path, proj)
        for e in es:
            entries_by_key.setdefault(e["item_key"], []).append(e)
        # Default metadata if not already known
        projects_meta.setdefault(proj, {
            "villa_type": "unknown", "floors": None,
            "plot_area":  None,      "gf_area": None,
        })
        print(f"  + {proj}: {len(es)} items")
    ref = _aggregate(entries_by_key, projects_meta, out_path)
    _print_summary(ref, out_path)


def mode_add(proj_path: str, out_path: str, meta: dict):
    """Incrementally add one project's BOQ to the DB."""
    proj_name = meta.get("project_name") or os.path.basename(proj_path.rstrip(os.sep))
    # Load existing DB
    if not os.path.exists(out_path):
        print(f"  No existing DB at {out_path} — run full rebuild first.")
        return
    with open(out_path, encoding='utf-8') as f:
        ref = json.load(f)
    projects_meta = ref.pop("_projects", {})
    # Re-expand existing per-item value lists into entry rows so we can re-aggregate
    entries_by_key: dict = {}
    for key, info in list(ref.items()):
        if key.startswith("_"):
            continue
        unit = info.get("unit", "")
        # We lost per-project linkage when we stored only the values list,
        # so synthesize "unknown_project" entries to preserve them
        for v in info.get("values", []):
            entries_by_key.setdefault(key, []).append({
                "item_key": key, "qty": v, "unit": unit,
                "project":  "_legacy", "desc": "",
            })
    # Add new project
    new_entries = parse_project_folder(proj_path, proj_name)
    for e in new_entries:
        entries_by_key.setdefault(e["item_key"], []).append(e)
    projects_meta["_legacy"] = projects_meta.get("_legacy", {
        "villa_type": "unknown", "floors": None,
        "plot_area":  None,      "gf_area": None,
        "_note": "pre-metadata projects",
    })
    projects_meta[proj_name] = {
        "villa_type": meta.get("villa_type", "unknown"),
        "floors":     meta.get("floors"),
        "plot_area":  meta.get("plot_area"),
        "gf_area":    meta.get("gf_area"),
    }
    ref2 = _aggregate(entries_by_key, projects_meta, out_path)
    _print_summary(ref2, out_path)
    print(f"\n  ✓ Added project '{proj_name}' ({len(new_entries)} items)")


# ── CLI ────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Build / extend reference_ranges.json")
    ap.add_argument("--base",    default=BASE,  help="Folder of project subfolders for full rebuild")
    ap.add_argument("--out",     default=r'C:\Users\basel\villa_qto\engine\reference_ranges.json',
                    help="Output reference_ranges.json path")
    ap.add_argument("--add",     help="Path to a single new project folder to ADD incrementally")
    ap.add_argument("--project-name", help="Override the project name (default: folder basename)")
    ap.add_argument("--villa-type",  default="unknown",
                    help="G | G+1 | G+2 | G+B+1 etc.")
    ap.add_argument("--floors",     type=int, help="Total structural levels (incl. roof)")
    ap.add_argument("--plot-area",  type=float)
    ap.add_argument("--gf-area",    type=float)
    args = ap.parse_args()

    if args.add:
        mode_add(args.add, args.out, {
            "project_name": args.project_name,
            "villa_type":   args.villa_type,
            "floors":       args.floors,
            "plot_area":    args.plot_area,
            "gf_area":      args.gf_area,
        })
    else:
        mode_rebuild(args.base, args.out)


if __name__ == "__main__":
    main()
